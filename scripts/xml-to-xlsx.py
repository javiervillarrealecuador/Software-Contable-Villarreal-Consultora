#!/usr/bin/env python3
"""
scripts/xml-to-xlsx.py
Parsea los XMLs descargados del SRI y genera un archivo Excel.

USO:
  python scripts/xml-to-xlsx.py <directorio_xmls> <archivo_salida.xlsx>

EJEMPLO:
  python scripts/xml-to-xlsx.py xmls-sri/0401200241001/2026-06 reporte.xlsx
  python scripts/xml-to-xlsx.py xmls-sri reporte-todos.xlsx    # todos los clientes

Estructura esperada del XML:
  <autorizacion>
    <estado>AUTORIZADO</estado>
    <numeroAutorizacion>...</numeroAutorizacion>
    <fechaAutorizacion>...</fechaAutorizacion>
    <comprobante><![CDATA[ ... XML del comprobante ... ]]></comprobante>
  </autorizacion>
"""

import os
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# Mapeo de códigos de tipo de comprobante del SRI
TIPO_COMPROBANTE = {
    "01": "Factura",
    "02": "Nota de Venta",
    "03": "Liquidación de Compra",
    "04": "Nota de Crédito",
    "05": "Nota de Débito",
    "06": "Guía de Remisión",
    "07": "Retención",
    "08": "Boleto",
}

# Mapeo de códigos de impuesto
CODIGO_IMPUESTO = {
    "0": "0%",
    "2": "12%",
    "3": "14%",
    "4": "15%",
    "6": "No Objeto",
    "7": "Exento",
}


def extract_text(el, tag, default=""):
    """Extrae texto de un subelemento XML."""
    child = el.find(tag)
    if child is not None and child.text:
        return child.text.strip()
    # Buscar recursivamente
    child = el.find(f".//{tag}")
    if child is not None and child.text:
        return child.text.strip()
    return default


def parse_comprobante_xml(xml_text):
    """Parsea el XML interno del comprobante y extrae datos relevantes."""
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return None

    # El root puede ser <factura>, <notaCredito>, <comprobanteRetencion>, etc.
    tag_root = root.tag.lower() if root.tag else ""

    # infoTributaria — común a todos los comprobantes
    info_trib = root.find("infoTributaria")
    if info_trib is None:
        return None

    data = {
        "tipo_comprobante_cod": extract_text(info_trib, "codDoc"),
        "ruc_emisor": extract_text(info_trib, "ruc"),
        "razon_social_emisor": extract_text(info_trib, "razonSocial"),
        "nombre_comercial": extract_text(info_trib, "nombreComercial"),
        "establecimiento": extract_text(info_trib, "estab"),
        "punto_emision": extract_text(info_trib, "ptoEmi"),
        "secuencial": extract_text(info_trib, "secuencial"),
        "clave_acceso": extract_text(info_trib, "claveAcceso"),
    }

    # Número completo: 001-001-000123456
    data["numero_comprobante"] = (
        f"{data['establecimiento']}-{data['punto_emision']}-{data['secuencial']}"
    )
    data["tipo_comprobante"] = TIPO_COMPROBANTE.get(
        data["tipo_comprobante_cod"], data["tipo_comprobante_cod"]
    )

    # Buscar la sección de info específica del comprobante
    # infoFactura, infoNotaCredito, infoCompRetencion, etc.
    info_tags = [
        "infoFactura",
        "infoNotaCredito",
        "infoNotaDebito",
        "infoCompRetencion",
        "infoLiquidacionCompra",
        "infoGuiaRemision",
    ]
    info_doc = None
    for tag in info_tags:
        info_doc = root.find(tag)
        if info_doc is not None:
            break

    if info_doc is not None:
        data["fecha_emision"] = extract_text(info_doc, "fechaEmision")
        data["ruc_comprador"] = extract_text(
            info_doc, "identificacionComprador",
            extract_text(info_doc, "identificacionSujetoRetenido", "")
        )
        data["razon_social_comprador"] = extract_text(
            info_doc, "razonSocialComprador",
            extract_text(info_doc, "razonSocialSujetoRetenido", "")
        )
        data["subtotal_sin_impuestos"] = to_float(
            extract_text(info_doc, "totalSinImpuestos", "0")
        )
        data["total_descuento"] = to_float(
            extract_text(info_doc, "totalDescuento", "0")
        )
        data["importe_total"] = to_float(
            extract_text(info_doc, "importeTotal",
                extract_text(info_doc, "valorTotal", "0"))
        )
        data["propina"] = to_float(extract_text(info_doc, "propina", "0"))
        data["moneda"] = extract_text(info_doc, "moneda", "DOLAR")

        # Impuestos (totalConImpuestos -> totalImpuesto)
        data["iva_0"] = 0.0
        data["iva_12"] = 0.0
        data["iva_15"] = 0.0
        data["base_no_iva"] = 0.0
        data["valor_iva"] = 0.0

        total_impuestos = info_doc.find("totalConImpuestos")
        if total_impuestos is not None:
            for imp in total_impuestos.findall("totalImpuesto"):
                codigo_porcentaje = extract_text(imp, "codigoPorcentaje")
                base = to_float(extract_text(imp, "baseImponible", "0"))
                valor = to_float(extract_text(imp, "valor", "0"))

                if codigo_porcentaje == "0":
                    data["iva_0"] = base
                elif codigo_porcentaje == "2":
                    data["iva_12"] = base
                    data["valor_iva"] += valor
                elif codigo_porcentaje in ("3", "4"):
                    data["iva_15"] = base
                    data["valor_iva"] += valor
                elif codigo_porcentaje in ("6", "7"):
                    data["base_no_iva"] = base
    else:
        data["fecha_emision"] = ""
        data["ruc_comprador"] = ""
        data["razon_social_comprador"] = ""
        data["subtotal_sin_impuestos"] = 0.0
        data["total_descuento"] = 0.0
        data["importe_total"] = 0.0
        data["propina"] = 0.0
        data["moneda"] = ""
        data["iva_0"] = 0.0
        data["iva_12"] = 0.0
        data["iva_15"] = 0.0
        data["base_no_iva"] = 0.0
        data["valor_iva"] = 0.0

    # Retenciones — si es comprobante de retención
    data["retenciones"] = []
    if tag_root in ("comprobanteretencion", "comprobanteRetencion"):
        impuestos_ret = root.find("impuestos") or root.find("docsSustento")
        if impuestos_ret is not None:
            for ret in impuestos_ret.findall("retencion"):
                data["retenciones"].append({
                    "codigo": extract_text(ret, "codigo"),
                    "codigo_retencion": extract_text(ret, "codigoRetencion"),
                    "base_imponible": to_float(extract_text(ret, "baseImponible", "0")),
                    "porcentaje": to_float(extract_text(ret, "porcentajeRetener", "0")),
                    "valor_retenido": to_float(extract_text(ret, "valorRetenido", "0")),
                })

    return data


def to_float(s):
    """Convierte string a float, retorna 0 si falla."""
    try:
        return float(s.replace(",", "."))
    except (ValueError, AttributeError):
        return 0.0


def find_xml_files(directory):
    """Encuentra todos los .xml en el directorio (recursivo)."""
    xml_files = []
    for root_dir, _, files in os.walk(directory):
        for f in files:
            if f.endswith(".xml"):
                xml_files.append(os.path.join(root_dir, f))
    return sorted(xml_files)


def parse_authorization_xml(filepath):
    """
    Parsea el archivo XML de autorización completo.
    Estructura: <autorizacion><comprobante>XML_REAL</comprobante></autorizacion>
    """
    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
    except ET.ParseError:
        return None

    # Si es el wrapper de autorización
    if root.tag == "autorizacion":
        estado = extract_text(root, "estado")
        num_aut = extract_text(root, "numeroAutorizacion")
        fecha_aut = extract_text(root, "fechaAutorizacion")

        comp_el = root.find("comprobante")
        if comp_el is not None and comp_el.text:
            comp_data = parse_comprobante_xml(comp_el.text.strip())
            if comp_data:
                comp_data["estado_autorizacion"] = estado
                comp_data["numero_autorizacion"] = num_aut
                comp_data["fecha_autorizacion"] = fecha_aut
                return comp_data
    else:
        # Es el XML directo del comprobante (sin wrapper)
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()
        comp_data = parse_comprobante_xml(content)
        if comp_data:
            comp_data["estado_autorizacion"] = ""
            comp_data["numero_autorizacion"] = ""
            comp_data["fecha_autorizacion"] = ""
            return comp_data

    return None


def create_excel(records, output_path):
    """Crea el archivo Excel con los datos parseados."""
    wb = Workbook()

    # ── Hoja 1: Comprobantes ──
    ws = wb.active
    ws.title = "Comprobantes"

    # Estilos
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=9)
    money_format = '#,##0.00'
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    headers = [
        ("Tipo", 14),
        ("Fecha Emisión", 13),
        ("RUC Emisor", 15),
        ("Razón Social Emisor", 35),
        ("Nro. Comprobante", 20),
        ("RUC Receptor", 15),
        ("Razón Social Receptor", 30),
        ("Subtotal 0%", 13),
        ("Subtotal IVA", 13),
        ("Base No IVA", 13),
        ("Descuento", 12),
        ("IVA", 12),
        ("Total", 13),
        ("Clave Acceso", 52),
        ("Nro. Autorización", 52),
        ("Fecha Autorización", 18),
        ("Estado", 12),
    ]

    # Escribir headers
    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Escribir datos
    for row_idx, r in enumerate(records, 2):
        values = [
            r.get("tipo_comprobante", ""),
            r.get("fecha_emision", ""),
            r.get("ruc_emisor", ""),
            r.get("razon_social_emisor", ""),
            r.get("numero_comprobante", ""),
            r.get("ruc_comprador", ""),
            r.get("razon_social_comprador", ""),
            r.get("iva_0", 0),
            r.get("iva_12", 0) + r.get("iva_15", 0),
            r.get("base_no_iva", 0),
            r.get("total_descuento", 0),
            r.get("valor_iva", 0),
            r.get("importe_total", 0),
            r.get("clave_acceso", ""),
            r.get("numero_autorizacion", ""),
            r.get("fecha_autorizacion", ""),
            r.get("estado_autorizacion", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = border
            # Formato monetario para columnas 8-13
            if 8 <= col <= 13 and isinstance(val, (int, float)):
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right")

    # Fila de totales
    if records:
        total_row = len(records) + 2
        ws.cell(row=total_row, column=6, value="TOTALES").font = Font(
            name="Arial", bold=True, size=10
        )
        for col in range(8, 14):  # columnas monetarias
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=total_row,
                column=col,
                value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
            )
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.number_format = money_format
            cell.border = border

    # Autofilter
    ws.auto_filter.ref = f"A1:Q{len(records) + 1}"

    # ── Hoja 2: Resumen por Emisor ──
    ws2 = wb.create_sheet("Resumen por Emisor")

    # Agrupar por emisor
    emisores = {}
    for r in records:
        ruc = r.get("ruc_emisor", "?")
        if ruc not in emisores:
            emisores[ruc] = {
                "razon_social": r.get("razon_social_emisor", ""),
                "count": 0,
                "subtotal": 0,
                "iva": 0,
                "total": 0,
            }
        emisores[ruc]["count"] += 1
        emisores[ruc]["subtotal"] += r.get("subtotal_sin_impuestos", 0)
        emisores[ruc]["iva"] += r.get("valor_iva", 0)
        emisores[ruc]["total"] += r.get("importe_total", 0)

    headers2 = [
        ("RUC Emisor", 15),
        ("Razón Social", 40),
        ("Comprobantes", 14),
        ("Subtotal", 14),
        ("IVA", 14),
        ("Total", 14),
    ]

    for col, (title, width) in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = width

    ws2.freeze_panes = "A2"

    for row_idx, (ruc, info) in enumerate(
        sorted(emisores.items(), key=lambda x: x[1]["total"], reverse=True), 2
    ):
        values2 = [
            ruc,
            info["razon_social"],
            info["count"],
            info["subtotal"],
            info["iva"],
            info["total"],
        ]
        for col, val in enumerate(values2, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = border
            if 4 <= col <= 6 and isinstance(val, (int, float)):
                cell.number_format = money_format

    wb.save(output_path)
    print(f"Excel guardado: {output_path}")
    print(f"  Comprobantes: {len(records)}")
    print(f"  Emisores únicos: {len(emisores)}")


def main():
    if len(sys.argv) < 3:
        print("USO: python scripts/xml-to-xlsx.py <directorio_xmls> <archivo_salida.xlsx>")
        print("EJEMPLO: python scripts/xml-to-xlsx.py xmls-sri/0401200241001/2026-06 reporte.xlsx")
        sys.exit(1)

    xml_dir = sys.argv[1]
    output = sys.argv[2]

    if not os.path.isdir(xml_dir):
        print(f"Error: {xml_dir} no es un directorio válido")
        sys.exit(1)

    xml_files = find_xml_files(xml_dir)
    print(f"Encontrados {len(xml_files)} archivos XML en {xml_dir}")

    if not xml_files:
        print("No hay XMLs para procesar.")
        sys.exit(0)

    records = []
    errors = 0
    for f in xml_files:
        data = parse_authorization_xml(f)
        if data:
            # Agregar info de la ruta (RUC del cliente / mes)
            parts = Path(f).parts
            # Intentar extraer RUC del path: xmls-sri/RUC/YYYY-MM/clave.xml
            data["_archivo"] = f
            records.append(data)
        else:
            errors += 1
            print(f"  ⚠ No se pudo parsear: {f}")

    print(f"Parseados: {len(records)} OK, {errors} errores")

    if records:
        create_excel(records, output)


if __name__ == "__main__":
    main()
